# -*- coding: utf-8 -*-
# @time     : 2019/1/9 0009 下午 15:50
# @Author   : yuxuan
# #file     : do_excel.py

import openpyxl
from common import project_path

# 初始化excel中列的case的属性
class Case:

   def __init__(self):
       self.case_id = None
       self.title = None
       self.url = None
       self.port_name = None
       self.data = None
       self.expected = None
       self.actual = None
       self.result = None



# 读取excel中测试用例
class DoExcel:

    # 打开excel放在初始化函数中
    def __init__(self, file_name):
        try:
            self.file_name = file_name
            self.db = openpyxl.load_workbook(filename=self.file_name)
        except FileNotFoundError as e:
            print('{0}file not found,please check file path'.format(self.file_name))
            raise e


    # 获取工作表中每一行case
    def get_cases(self, sheet_name):
        sheet = self.db[sheet_name]
        max_row = sheet.max_row
        cases = []
        for i in range(2,max_row+1):
            case = Case()
            case.case_id = sheet.cell(i, 1).value
            case.title = sheet.cell(i, 2).value
            case.url = sheet.cell(i, 3).value
            case.port_name = sheet.cell(i, 4).value
            case.data = sheet.cell(i, 5).value
            case.expected = sheet.cell(i, 6).value
            case.actual = sheet.cell(i, 7).value
            case.result = sheet.cell(i, 8).value
            cases.append(case)
        return cases


    def bace_write_by_case_id(self, sheet_name, case_id, actual, result):

        sheet = self.db[sheet_name]
        max_row = sheet.max_row
        for r in range(2, max_row+1):
            if sheet.cell(r, 1).value == case_id:    # 判断传入的case_id行是否和读取工作表中获取行相同
                sheet.cell(r, 7).value = actual
                sheet.cell(r, 8).value = result
                self.db.save(filename=self.file_name)
                break


if __name__ == '__main__':
    do_excel = DoExcel(project_path.data_path)
    do_excel.bace_write_by_case_id('sendMcode', 1, '用户IP不能为空', 'success')
























